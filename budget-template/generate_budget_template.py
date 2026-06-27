"""
Mochi Collective — Hackathon / Innovation Sprint Budget Template generator.

Produces a formula-driven Excel workbook that lets you:
  1. Estimate cost for a 20-pax and a 50-pax event.
  2. Track budget vs actual on a live project, auto-flagging any line item
     that runs more than +20% over its budget.
  3. Capture the circumstances behind variances, the impact of high-value
     spend, and the way forward on over-budget items.

Re-run this script to regenerate the .xlsx:
    python3 generate_budget_template.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

OUTPUT = "Hackathon_Innovation_Sprint_Budget_Template.xlsx"

# ----------------------------------------------------------------------------
# Brand palette (kept close to Mochi Collective's soft, warm aesthetic)
# ----------------------------------------------------------------------------
INK = "1F2933"          # near-black text
PLUM = "5B3A56"         # deep accent for headers
MOCHI = "E8DCC8"        # warm beige band
BLUSH = "F4E3DA"        # soft section band
MINT = "DCEAD9"         # positive / under-budget
ALERT = "F6C6C6"        # over-budget flag
ALERT_TXT = "9B1C1C"
SOFT_GREY = "F3F1ED"
LINE = "D9D2C7"
WHITE = "FFFFFF"

thin = Side(style="thin", color=LINE)
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
border_bottom = Border(bottom=Side(style="thin", color=PLUM))

title_font = Font(name="Calibri", size=18, bold=True, color=WHITE)
sub_font = Font(name="Calibri", size=10, italic=True, color="FFFFFF")
section_font = Font(name="Calibri", size=12, bold=True, color=WHITE)
header_font = Font(name="Calibri", size=10, bold=True, color=WHITE)
label_font = Font(name="Calibri", size=10, bold=True, color=INK)
body_font = Font(name="Calibri", size=10, color=INK)
muted_font = Font(name="Calibri", size=9, italic=True, color="6B7280")
cat_font = Font(name="Calibri", size=10, bold=True, color=PLUM)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
right = Alignment(horizontal="right", vertical="center")

plum_fill = PatternFill("solid", fgColor=PLUM)
mochi_fill = PatternFill("solid", fgColor=MOCHI)
blush_fill = PatternFill("solid", fgColor=BLUSH)
grey_fill = PatternFill("solid", fgColor=SOFT_GREY)
white_fill = PatternFill("solid", fgColor=WHITE)
mint_fill = PatternFill("solid", fgColor=MINT)
alert_fill = PatternFill("solid", fgColor=ALERT)

CURRENCY_FMT = '#,##0.00'
PCT_FMT = '0.0%'

# ----------------------------------------------------------------------------
# Line item data: (Category, Line item, Description / assumption, est20, est50)
# Amounts are ILLUSTRATIVE planning placeholders — overwrite with your quotes.
# Fixed-ish costs scale gently; per-head costs scale with headcount.
# ----------------------------------------------------------------------------
GTH = "  (good-to-have)"   # marker appended to client-rapport / optional items

ITEMS = [
    # --- Venue & Facilities -------------------------------------------------
    ("Venue & Facilities", "Venue hire (full day)", "Main space, full-day rate", 1500, 3000),
    ("Venue & Facilities", "Breakout rooms", "Team rooms for parallel work", 400, 1000),
    ("Venue & Facilities", "AV equipment rental", "Projector, screens, mics, PA", 600, 1200),
    ("Venue & Facilities", "Wifi / bandwidth upgrade", "Guaranteed uplink for demos", 200, 500),
    ("Venue & Facilities", "Furniture / layout reconfig", "Cabaret/cluster setup, tear-down", 150, 400),
    ("Venue & Facilities", "Power / charging stations", "Extension runs, charging carts", 100, 250),
    ("Venue & Facilities", "Cleaning / janitorial", "Mid-event + post-event", 150, 350),

    # --- Technology & Tools -------------------------------------------------
    ("Technology & Tools", "Cloud / API credits", "Prototyping compute & API usage", 300, 750),
    ("Technology & Tools", "Software / SaaS licenses", "Short-term tooling for the sprint", 200, 500),
    ("Technology & Tools", "Collaboration tools", "Miro / FigJam / Slack workspace", 150, 350),
    ("Technology & Tools", "Hardware kits / components", "IoT / maker kits (if applicable)", 0, 0),
    ("Technology & Tools", "Device rental", "Loaner laptops / tablets (if needed)", 0, 0),
    ("Technology & Tools", "Printing & signage", "Wayfinding, schedule, brief packs", 150, 300),

    # --- Facilitation & Talent ---------------------------------------------
    ("Facilitation & Talent", "Lead facilitator / host", "Design + delivery day rate", 1200, 1500),
    ("Facilitation & Talent", "Co-facilitators / mentors", "Floor coaching, 1 per ~10 pax", 800, 2000),
    ("Facilitation & Talent", "SMEs / guest speakers", "Domain inputs, kick-off talk", 500, 1000),
    ("Facilitation & Talent", "Judges (honorarium)", "Panel for final pitches", 300, 600),
    ("Facilitation & Talent", "Technical support staff", "On-call IT / runner", 300, 700),
    ("Facilitation & Talent", "Photographer / videographer", "Coverage across the day", 600, 900),

    # --- Food & Beverage ----------------------------------------------------
    ("Food & Beverage", "Breakfast / arrival pastries", "Per head", 200, 500),
    ("Food & Beverage", "Lunch", "Per head, catered", 400, 1000),
    ("Food & Beverage", "Snacks & energy bites", "Afternoon refuel", 200, 450),
    ("Food & Beverage", "Coffee / tea / beverages", "All-day station", 250, 550),
    ("Food & Beverage", "Dietary requirements buffer", "Halal / vegan / allergen", 100, 250),
    ("Food & Beverage", "Closing reception F&B" + GTH, "Celebratory wrap", 400, 1000),

    # --- Participant Experience --------------------------------------------
    ("Participant Experience", "Welcome kit / swag", "Notebook, sticker, essentials", 400, 900),
    ("Participant Experience", "Name badges & lanyards", "Printed + holders", 80, 180),
    ("Participant Experience", "Stationery & materials", "Post-its, markers, paper, dots", 150, 350),
    ("Participant Experience", "Prizes / awards", "Winning + runner-up teams", 1000, 2000),
    ("Participant Experience", "Certificates", "Printed / digital", 50, 120),

    # --- Marketing & Communications ----------------------------------------
    ("Marketing & Communications", "Event branding & design", "Identity, deck, templates", 500, 800),
    ("Marketing & Communications", "Printed collateral & banners", "Backdrop, standees", 200, 400),
    ("Marketing & Communications", "Photo / video editing", "Highlight reel, recap edit", 500, 900),
    ("Marketing & Communications", "Social media / promotion", "Pre / during / post posts", 200, 500),

    # --- Logistics & Operations --------------------------------------------
    ("Logistics & Operations", "Transportation / parking", "Validations, shuttles", 150, 350),
    ("Logistics & Operations", "Speaker / facilitator travel", "Transport for talent", 200, 500),
    ("Logistics & Operations", "Accommodation", "Out-of-town guests (if needed)", 0, 0),
    ("Logistics & Operations", "Shipping / courier", "Kit & equipment movement", 100, 200),
    ("Logistics & Operations", "Insurance / permits", "Event cover, licences", 200, 400),
    ("Logistics & Operations", "First aid / safety", "Kit, on-site cover", 80, 150),

    # --- Client Rapport / Good-to-haves ------------------------------------
    ("Client Rapport / Good-to-haves", "Premium welcome (barista cart)" + GTH, "Wow-moment on arrival", 400, 800),
    ("Client Rapport / Good-to-haves", "Premium branded swag upgrade" + GTH, "Apparel / quality merch", 300, 750),
    ("Client Rapport / Good-to-haves", "Networking / closing celebration" + GTH, "Drinks + canapes", 500, 1200),
    ("Client Rapport / Good-to-haves", "Executive gift / token" + GTH, "For client stakeholders", 200, 400),
    ("Client Rapport / Good-to-haves", "Cinematic highlight video" + GTH, "Polished hero film", 800, 1500),
    ("Client Rapport / Good-to-haves", "Post-event impact report & deck" + GTH, "Outcomes + ROI narrative", 500, 800),
    ("Client Rapport / Good-to-haves", "Wellness corner" + GTH, "Massage / meditation break", 300, 600),
]

CATEGORIES = []
for cat, *_ in ITEMS:
    if cat not in CATEGORIES:
        CATEGORIES.append(cat)


def style_block(ws, cell_range, font=None, fill=None, align=None, border=None, fmt=None):
    for row in ws[cell_range]:
        for c in row:
            if font:
                c.font = font
            if fill:
                c.fill = fill
            if align:
                c.alignment = align
            if border:
                c.border = border
            if fmt:
                c.number_format = fmt


# ===========================================================================
wb = Workbook()

# ---------------------------------------------------------------------------
# SHEET 1 — How to use
# ---------------------------------------------------------------------------
ws0 = wb.active
ws0.title = "How to use"
ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 3
ws0.column_dimensions["B"].width = 104

ws0.merge_cells("B2:B2")
ws0["B2"] = "Hackathon / Innovation Sprint — Budget Template"
ws0["B2"].font = Font(name="Calibri", size=20, bold=True, color=PLUM)

guide = [
    ("", ""),
    ("WHAT THIS WORKBOOK DOES", "h"),
    ("Plan the cost of a 20-pax or 50-pax hackathon / innovation sprint, then track budget "
     "vs. actual spend on a live project — with automatic flags and a structured place to "
     "explain the big swings.", "p"),
    ("", ""),
    ("THE TABS", "h"),
    ("1.  Budget vs Actual  —  the working sheet. Pick your event size, set budgets, log "
     "actuals. Any line item that runs >20% over budget is flagged red and a reason becomes "
     "required.", "p"),
    ("2.  Dashboard  —  auto-calculated roll-up by category, totals, and the list of flagged "
     "over-budget items. Nothing to fill in here.", "p"),
    ("3.  Impact & Opportunities  —  log the spend that punched above its weight and the "
     "future opportunities it opened.", "p"),
    ("4.  Over-Budget Deep Dive  —  for each flagged item: the reason, the impact it actually "
     "delivered, and the way forward.", "p"),
    ("", ""),
    ("HOW TO USE IT", "h"),
    ("Step 1  —  On 'Budget vs Actual', fill the project header and choose Event size (20 or "
     "50) from the dropdown. The Budgeted column auto-fills from the matching estimate; "
     "overwrite any cell with your real quote.", "p"),
    ("Step 2  —  As money is committed, enter the Actual for each line. Variance ($ and %) and "
     "the over-budget flag calculate automatically.", "p"),
    ("Step 3  —  Whenever 'Flag' shows ⚠ OVER, write the circumstance in the Reason column. "
     "This is the audit trail for spend that ran >20% hot.", "p"),
    ("Step 4  —  After the event, complete 'Impact & Opportunities' and 'Over-Budget Deep "
     "Dive' to close the loop on value and learnings.", "p"),
    ("", ""),
    ("CONVENTIONS", "h"),
    ("•  Amounts shipped in the template are ILLUSTRATIVE planning placeholders — replace them "
     "with your own quotes.", "p"),
    ("•  Set your currency once in the header cell on 'Budget vs Actual'; it is shown on the "
     "Dashboard too.", "p"),
    ("•  Items tagged '(good-to-have)' are optional client-rapport extras — keep, cut, or cost "
     "as the relationship warrants.", "p"),
    ("•  The +20% threshold is the trigger for an explanation. You can change it in the header "
     "(Variance threshold cell).", "p"),
]
r = 3
for text, kind in guide:
    cell = ws0.cell(row=r, column=2, value=text)
    if kind == "h":
        cell.font = Font(name="Calibri", size=12, bold=True, color=INK)
        cell.fill = mochi_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws0.row_dimensions[r].height = 22
    elif kind == "p":
        cell.font = body_font
        cell.alignment = left_top
        ws0.row_dimensions[r].height = 30 if len(text) > 95 else 18
    r += 1

# ---------------------------------------------------------------------------
# SHEET 2 — Budget vs Actual (the working sheet)
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Budget vs Actual")
ws.sheet_view.showGridLines = False

# Columns: A Category | B Line item | C Description | D Est20 | E Est50
#          F Budgeted | G Actual | H Var$ | I Var% | J Flag | K Reason
widths = [22, 30, 30, 13, 13, 14, 14, 13, 11, 13, 40]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# --- Title band ---
ws.merge_cells("A1:K1")
ws["A1"] = "BUDGET vs ACTUAL  —  Hackathon / Innovation Sprint"
ws["A1"].font = title_font
ws["A1"].fill = plum_fill
ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 30

# --- Project header block (rows 2-4) ---
hdr = {
    "A2": "Project / Client:", "C2": "",            # free text
    "F2": "Event date:", "G2": "",
    "A3": "Event size (pax):", "B3": 20,
    "C3": "Currency:", "D3": "SGD",
    "F3": "Variance threshold:", "G3": 0.20,
    "A4": "Contingency %:", "B4": 0.10,
    "C4": "Prepared by:", "D4": "",
}
for ref, val in hdr.items():
    ws[ref] = val
for ref in ["A2", "F2", "A3", "C3", "F3", "A4", "C4"]:
    ws[ref].font = label_font
    ws[ref].alignment = right
for ref in ["C2", "G2", "B3", "D3", "G3", "B4", "D4"]:
    ws[ref].fill = grey_fill
    ws[ref].font = body_font
    ws[ref].alignment = center
    ws[ref].border = border_all
ws.merge_cells("C2:E2")   # project name
ws["C2"].alignment = left
ws["G2"].number_format = "yyyy-mm-dd"
ws["G3"].number_format = PCT_FMT
ws["B4"].number_format = PCT_FMT
ws["B3"].number_format = "0"
for rr in (2, 3, 4):
    ws.row_dimensions[rr].height = 20

# Event size dropdown (20 / 50)
dv_pax = DataValidation(type="list", formula1='"20,50"', allow_blank=False)
ws.add_data_validation(dv_pax)
dv_pax.add(ws["B3"])

# --- Table header (row 6) ---
HEAD_ROW = 6
headers = [
    "Category", "Line item", "Description / assumption",
    "Est. 20 pax", "Est. 50 pax", "Budgeted", "Actual",
    "Variance", "Variance %", "Flag", "Reason / circumstances (required if flagged)",
]
for col, text in enumerate(headers, start=1):
    c = ws.cell(row=HEAD_ROW, column=col, value=text)
    c.font = header_font
    c.fill = plum_fill
    c.alignment = center
    c.border = border_all
ws.row_dimensions[HEAD_ROW].height = 30
ws.freeze_panes = "A7"

# --- Data rows ---
first_data = HEAD_ROW + 1
row = first_data
data_rows = []
for cat, item, desc, e20, e50 in ITEMS:
    # Category cell: only label on first row of each category group
    ws.cell(row=row, column=1, value=cat).font = cat_font
    ws.cell(row=row, column=1).alignment = left_top
    ws.cell(row=row, column=2, value=item).font = body_font
    ws.cell(row=row, column=2).alignment = left_top
    ws.cell(row=row, column=3, value=desc).font = muted_font
    ws.cell(row=row, column=3).alignment = left_top
    ws.cell(row=row, column=4, value=e20).number_format = CURRENCY_FMT
    ws.cell(row=row, column=5, value=e50).number_format = CURRENCY_FMT
    # Budgeted auto-fills from the matching estimate based on the pax selector
    ws.cell(row=row, column=6,
            value=f"=IF($B$3=50,E{row},D{row})").number_format = CURRENCY_FMT
    ws.cell(row=row, column=7, value=None).number_format = CURRENCY_FMT       # Actual (input)
    ws.cell(row=row, column=8, value=f"=IF(G{row}=\"\",\"\",G{row}-F{row})").number_format = CURRENCY_FMT
    ws.cell(row=row, column=9,
            value=f"=IF(OR(G{row}=\"\",F{row}=0),\"\",(G{row}-F{row})/F{row})").number_format = PCT_FMT
    ws.cell(row=row, column=10,
            value=f"=IF(I{row}=\"\",\"\",IF(I{row}>$G$3,\"⚠ OVER\",IF(I{row}<-$G$3,\"under\",\"ok\")))")
    ws.cell(row=row, column=11, value=None)   # Reason (input)
    for col in range(1, 12):
        ws.cell(row=row, column=col).border = border_all
    for col in (4, 5, 6, 7, 8):
        ws.cell(row=row, column=col).alignment = right
    for col in (9, 10):
        ws.cell(row=row, column=col).alignment = center
    ws.cell(row=row, column=11).alignment = left_top
    # input cells get a soft fill so users see where to type
    ws.cell(row=row, column=7).fill = grey_fill
    ws.cell(row=row, column=11).fill = grey_fill
    ws.row_dimensions[row].height = 26
    data_rows.append(row)
    row += 1

last_data = row - 1

# --- Totals block ---
row += 0
sub_row = row
ws.cell(row=sub_row, column=2, value="SUBTOTAL (line items)").font = label_font
ws.cell(row=sub_row, column=4, value=f"=SUM(D{first_data}:D{last_data})")
ws.cell(row=sub_row, column=5, value=f"=SUM(E{first_data}:E{last_data})")
ws.cell(row=sub_row, column=6, value=f"=SUM(F{first_data}:F{last_data})")
ws.cell(row=sub_row, column=7, value=f"=SUM(G{first_data}:G{last_data})")
ws.cell(row=sub_row, column=8, value=f"=IF(SUM(G{first_data}:G{last_data})=0,\"\",G{sub_row}-F{sub_row})")
ws.cell(row=sub_row, column=9, value=f"=IF(OR(G{sub_row}=0,F{sub_row}=0),\"\",(G{sub_row}-F{sub_row})/F{sub_row})")

cont_row = sub_row + 1
ws.cell(row=cont_row, column=2, value="Contingency").font = label_font
ws.cell(row=cont_row, column=4, value=f"=D{sub_row}*$B$4")
ws.cell(row=cont_row, column=5, value=f"=E{sub_row}*$B$4")
ws.cell(row=cont_row, column=6, value=f"=F{sub_row}*$B$4")
ws.cell(row=cont_row, column=7, value=None).fill = grey_fill   # actual contingency drawn (input)

grand_row = cont_row + 1
ws.cell(row=grand_row, column=2, value="GRAND TOTAL").font = Font(size=11, bold=True, color=WHITE)
ws.cell(row=grand_row, column=4, value=f"=D{sub_row}+D{cont_row}")
ws.cell(row=grand_row, column=5, value=f"=E{sub_row}+E{cont_row}")
ws.cell(row=grand_row, column=6, value=f"=F{sub_row}+F{cont_row}")
ws.cell(row=grand_row, column=7, value=f"=SUM(G{first_data}:G{last_data})+N(G{cont_row})")
ws.cell(row=grand_row, column=8, value=f"=IF(G{grand_row}=0,\"\",G{grand_row}-F{grand_row})")
ws.cell(row=grand_row, column=9, value=f"=IF(OR(G{grand_row}=0,F{grand_row}=0),\"\",(G{grand_row}-F{grand_row})/F{grand_row})")

for rr in (sub_row, cont_row, grand_row):
    for col in range(1, 12):
        c = ws.cell(row=rr, column=col)
        c.border = border_all
        if col in (4, 5, 6, 7, 8):
            c.number_format = CURRENCY_FMT
            c.alignment = right
        if col == 9:
            c.number_format = PCT_FMT
            c.alignment = center
    ws.row_dimensions[rr].height = 22

style_block(ws, f"A{sub_row}:K{cont_row}", fill=mochi_fill)
style_block(ws, f"A{grand_row}:K{grand_row}", fill=plum_fill)
ws.cell(row=grand_row, column=4).font = Font(size=11, bold=True, color=WHITE)
ws.cell(row=grand_row, column=5).font = Font(size=11, bold=True, color=WHITE)
ws.cell(row=grand_row, column=6).font = Font(size=11, bold=True, color=WHITE)
ws.cell(row=grand_row, column=7).font = Font(size=11, bold=True, color=WHITE)

# --- Conditional formatting on the data range ---
flag_range = f"J{first_data}:J{last_data}"
ws.conditional_formatting.add(
    flag_range,
    CellIsRule(operator="equal", formula=['"⚠ OVER"'], fill=alert_fill,
               font=Font(bold=True, color=ALERT_TXT)))
ws.conditional_formatting.add(
    flag_range,
    CellIsRule(operator="equal", formula=['"under"'], fill=mint_fill))
# Highlight whole row's variance% when over threshold
var_range = f"I{first_data}:I{last_data}"
ws.conditional_formatting.add(
    var_range,
    FormulaRule(formula=[f"AND($I{first_data}<>\"\",$I{first_data}>$G$3)"],
                fill=alert_fill, font=Font(bold=True, color=ALERT_TXT)))
# Reason cell turns amber if flagged but empty (a prompt to fill it in)
reason_range = f"K{first_data}:K{last_data}"
ws.conditional_formatting.add(
    reason_range,
    FormulaRule(formula=[f"AND($J{first_data}=\"⚠ OVER\",$K{first_data}=\"\")"],
                fill=PatternFill("solid", fgColor="FCE8C8")))

# Note under the table
note_row = grand_row + 2
ws.merge_cells(f"A{note_row}:K{note_row}")
ws.cell(row=note_row, column=1,
        value="Tip: the Budgeted column auto-fills from the 20- or 50-pax estimate based on the "
              "Event size selector — overwrite any cell with your actual quote. A line flags ⚠ OVER "
              "once Actual exceeds Budgeted by more than the Variance threshold; document the why in "
              "the Reason column, then expand on it in the Over-Budget Deep Dive tab.")
ws.cell(row=note_row, column=1).font = muted_font
ws.cell(row=note_row, column=1).alignment = left_top
ws.row_dimensions[note_row].height = 42

# ---------------------------------------------------------------------------
# SHEET 3 — Dashboard (auto roll-up)
# ---------------------------------------------------------------------------
wd = wb.create_sheet("Dashboard")
wd.sheet_view.showGridLines = False
for i, w in enumerate([3, 30, 16, 16, 16, 14, 12], start=1):
    wd.column_dimensions[get_column_letter(i)].width = w

wd.merge_cells("B1:G1")
wd["B1"] = "DASHBOARD  —  Budget vs Actual roll-up"
wd["B1"].font = title_font
wd["B1"].fill = plum_fill
wd["B1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
wd.row_dimensions[1].height = 28

wd["B2"] = "Auto-calculated from the 'Budget vs Actual' tab — nothing to enter here."
wd["B2"].font = muted_font

SRC = "'Budget vs Actual'"
drow = 4
dcols = ["Category", "Budgeted", "Actual", "Variance", "Variance %", "Flag"]
for col, text in enumerate(dcols, start=2):
    c = wd.cell(row=drow, column=col, value=text)
    c.font = header_font
    c.fill = plum_fill
    c.alignment = center
    c.border = border_all
wd.row_dimensions[drow].height = 24

cat_first = drow + 1
rr = cat_first
for cat in CATEGORIES:
    wd.cell(row=rr, column=2, value=cat).font = body_font
    wd.cell(row=rr, column=2).alignment = left
    wd.cell(row=rr, column=3,
            value=f"=SUMIF({SRC}!$A${first_data}:$A${last_data},$B{rr},{SRC}!$F${first_data}:$F${last_data})")
    wd.cell(row=rr, column=4,
            value=f"=SUMIF({SRC}!$A${first_data}:$A${last_data},$B{rr},{SRC}!$G${first_data}:$G${last_data})")
    wd.cell(row=rr, column=5, value=f"=D{rr}-C{rr}")
    wd.cell(row=rr, column=6, value=f"=IF(C{rr}=0,\"\",(D{rr}-C{rr})/C{rr})")
    wd.cell(row=rr, column=7,
            value=f"=IF(C{rr}=0,\"\",IF((D{rr}-C{rr})/C{rr}>{SRC}!$G$3,\"⚠ OVER\","
                  f"IF((D{rr}-C{rr})/C{rr}<-{SRC}!$G$3,\"under\",\"ok\")))")
    for col in range(2, 8):
        wd.cell(row=rr, column=col).border = border_all
    for col in (3, 4, 5):
        wd.cell(row=rr, column=col).number_format = CURRENCY_FMT
        wd.cell(row=rr, column=col).alignment = right
    wd.cell(row=rr, column=6).number_format = PCT_FMT
    wd.cell(row=rr, column=6).alignment = center
    wd.cell(row=rr, column=7).alignment = center
    rr += 1
cat_last = rr - 1

# Totals row
wd.cell(row=rr, column=2, value="TOTAL").font = Font(size=11, bold=True, color=WHITE)
wd.cell(row=rr, column=3, value=f"=SUM(C{cat_first}:C{cat_last})")
wd.cell(row=rr, column=4, value=f"=SUM(D{cat_first}:D{cat_last})")
wd.cell(row=rr, column=5, value=f"=D{rr}-C{rr}")
wd.cell(row=rr, column=6, value=f"=IF(C{rr}=0,\"\",(D{rr}-C{rr})/C{rr})")
for col in range(2, 8):
    c = wd.cell(row=rr, column=col)
    c.fill = plum_fill
    c.border = border_all
    c.font = Font(size=11, bold=True, color=WHITE)
    if col in (3, 4, 5):
        c.number_format = CURRENCY_FMT
        c.alignment = right
    if col == 6:
        c.number_format = PCT_FMT
        c.alignment = center
total_row = rr

# stripe categories
for i, rr2 in enumerate(range(cat_first, cat_last + 1)):
    if i % 2 == 0:
        style_block(wd, f"B{rr2}:G{rr2}", fill=grey_fill)

# CF on dashboard flag/variance
wd.conditional_formatting.add(
    f"G{cat_first}:G{cat_last}",
    CellIsRule(operator="equal", formula=['"⚠ OVER"'], fill=alert_fill,
               font=Font(bold=True, color=ALERT_TXT)))
wd.conditional_formatting.add(
    f"F{cat_first}:F{cat_last}",
    CellIsRule(operator="greaterThan", formula=[f"{SRC}!$G$3"], fill=alert_fill,
               font=Font(bold=True, color=ALERT_TXT)))

# Headline KPI cards
k = total_row + 3
wd.cell(row=k, column=2, value="HEADLINE").font = Font(size=12, bold=True, color=PLUM)
cards = [
    ("Total budgeted", f"=C{total_row}", CURRENCY_FMT),
    ("Total actual", f"=D{total_row}", CURRENCY_FMT),
    ("Total variance", f"=D{total_row}-C{total_row}", CURRENCY_FMT),
    ("Variance %", f"=IF(C{total_row}=0,\"\",(D{total_row}-C{total_row})/C{total_row})", PCT_FMT),
    ("Line items flagged >20% over",
     f"=COUNTIF({SRC}!$J${first_data}:$J${last_data},\"⚠ OVER\")", "0"),
]
kr = k + 1
for label, formula, fmt in cards:
    wd.cell(row=kr, column=2, value=label).font = label_font
    wd.cell(row=kr, column=2).fill = mochi_fill
    wd.cell(row=kr, column=2).border = border_all
    wd.cell(row=kr, column=2).alignment = left
    c = wd.cell(row=kr, column=3, value=formula)
    c.number_format = fmt
    c.font = Font(size=11, bold=True, color=INK)
    c.fill = white_fill
    c.border = border_all
    c.alignment = right
    kr += 1


# ---------------------------------------------------------------------------
# Helper to build a "form table" section sheet
# ---------------------------------------------------------------------------
def build_section(ws, title, intro, columns, n_rows, color_band=BLUSH):
    ws.sheet_view.showGridLines = False
    total_w = 0
    for i, (_, w) in enumerate(columns, start=2):
        ws.column_dimensions[get_column_letter(i)].width = w
        total_w += w
    ws.column_dimensions["A"].width = 3
    last_col = get_column_letter(1 + len(columns))
    ws.merge_cells(f"B1:{last_col}1")
    ws["B1"] = title
    ws["B1"].font = title_font
    ws["B1"].fill = plum_fill
    ws["B1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28
    ws.merge_cells(f"B2:{last_col}2")
    ws["B2"] = intro
    ws["B2"].font = muted_font
    ws["B2"].alignment = left_top
    ws.row_dimensions[2].height = 30
    head = 4
    for col, (name, _) in enumerate(columns, start=2):
        c = ws.cell(row=head, column=col, value=name)
        c.font = header_font
        c.fill = plum_fill
        c.alignment = center
        c.border = border_all
    ws.row_dimensions[head].height = 30
    for ri in range(head + 1, head + 1 + n_rows):
        for col in range(2, 2 + len(columns)):
            c = ws.cell(row=ri, column=col)
            c.border = border_all
            c.alignment = left_top
            c.fill = white_fill if (ri - head) % 2 else grey_fill
        ws.row_dimensions[ri].height = 40
    return head


# ---------------------------------------------------------------------------
# SHEET 4 — Impact & Opportunities
# ---------------------------------------------------------------------------
wi = wb.create_sheet("Impact & Opportunities")
build_section(
    wi,
    "IMPACT & OPPORTUNITIES",
    "Capture the spend that delivered outsized value — the line items worth protecting or "
    "repeating — and the future opportunities each one opened (new leads, referrals, content, "
    "partnerships, capability).",
    [
        ("Spend / element", 26),
        ("Amount spent", 14),
        ("Why it was impactful (the value it created)", 46),
        ("Evidence / signal (feedback, metric, quote)", 38),
        ("Future opportunity it opened", 40),
        ("Owner / next step", 24),
    ],
    n_rows=12,
)
wi["C5"].number_format = CURRENCY_FMT  # amount column under header at row5 -> col C
# amount column is the 2nd column => column C
for rr in range(5, 17):
    wi.cell(row=rr, column=3).number_format = CURRENCY_FMT
    wi.cell(row=rr, column=3).alignment = right

# ---------------------------------------------------------------------------
# SHEET 5 — Over-Budget Deep Dive
# ---------------------------------------------------------------------------
wo = wb.create_sheet("Over-Budget Deep Dive")
build_section(
    wo,
    "OVER-BUDGET DEEP DIVE",
    "One row per line item flagged >20% over budget (see the Flag column on 'Budget vs Actual'). "
    "Diagnose the reason, assess the impact the extra spend actually delivered, and decide the "
    "way forward so the next event is better estimated.",
    [
        ("Line item", 24),
        ("Category", 20),
        ("Budgeted", 13),
        ("Actual", 13),
        ("Variance %", 12),
        ("Reason(s) for overspend", 40),
        ("Impact it actually gave", 36),
        ("Worth it? (Y/N)", 12),
        ("Way forward / action for next time", 40),
    ],
    n_rows=12,
)
# numeric formats for budget/actual/var% columns (D, E, F)
for rr in range(5, 17):
    wo.cell(row=rr, column=4).number_format = CURRENCY_FMT
    wo.cell(row=rr, column=4).alignment = right
    wo.cell(row=rr, column=5).number_format = CURRENCY_FMT
    wo.cell(row=rr, column=5).alignment = right
    wo.cell(row=rr, column=6).number_format = PCT_FMT
    wo.cell(row=rr, column=6).alignment = center
    wo.cell(row=rr, column=9).alignment = center
# Worth it dropdown
dv_worth = DataValidation(type="list", formula1='"Yes,No,Partly"', allow_blank=True)
wo.add_data_validation(dv_worth)
dv_worth.add(f"I5:I16")

# ---------------------------------------------------------------------------
wb.save(OUTPUT)
print(f"Wrote {OUTPUT}")
print(f"  line items: {len(ITEMS)}  |  categories: {len(CATEGORIES)}")
print(f"  data rows {first_data}-{last_data}, grand total row {grand_row}")
